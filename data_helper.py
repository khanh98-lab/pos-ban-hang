# data_helper.py
import pandas as pd
import sqlite3
import os
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

DB_FILE = 'data/pos_database.db'

def init_db():
    os.makedirs('data', exist_ok=True)
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS don_hang (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            so_don_hang TEXT,
            ngay_tao TEXT,
            chi_nhanh TEXT,
            ten_nv TEXT,
            ma_kh TEXT,
            ten_kh TEXT,
            ten_nguoi_mua TEXT,
            ma_so_thue TEXT,
            dia_chi TEXT,
            ma_vt TEXT,
            ten_vt TEXT,
            so_luong REAL,
            don_gia REAL,
            thanh_tien REAL,
            hinh_thuc_tt TEXT,
            xuat_hd TEXT,
            tien_tm REAL,
            tien_ck REAL,
            tien_no REAL,
            nha_may TEXT
        )
    ''')
    conn.commit()
    conn.close()

def luu_don_hang_danh_sach(chi_nhanh, ten_nv, ma_kh, ten_kh, ten_nguoi_mua, mst, dia_chi, danh_sach_hang, hinh_thuc_tt, xuat_hd, tien_tm, tien_ck, tien_no):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    so_don = f"DH_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    ngay_tao = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    val_ten_kh = str(ten_kh).strip() if ten_kh else ""
    final_ten_kh = "" if val_ten_kh.lower() == 'khách lẻ' else val_ten_kh

    val_ten_nguoi_mua = str(ten_nguoi_mua).strip() if ten_nguoi_mua else ""
    final_ten_nguoi_mua = "" if val_ten_nguoi_mua.lower() == 'khách lẻ' else val_ten_nguoi_mua

    for item in danh_sach_hang:
        sl = float(item.get('so_luong', 0))
        dg = float(item.get('don_gia', 0))
        tt = float(item.get('thanh_tien', sl * dg))
        
        cursor.execute('''
            INSERT INTO don_hang (
                so_don_hang, ngay_tao, chi_nhanh, ten_nv, ma_kh, ten_kh, ten_nguoi_mua, ma_so_thue, dia_chi,
                ma_vt, ten_vt, so_luong, don_gia, thanh_tien, hinh_thuc_tt, xuat_hd, tien_tm, tien_ck, tien_no, nha_may
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            so_don, ngay_tao, chi_nhanh, ten_nv, ma_kh, final_ten_kh, final_ten_nguoi_mua, mst, dia_chi,
            item.get('ma_vt', ''), item.get('ten_vt', ''), sl, dg, tt,
            hinh_thuc_tt, xuat_hd, float(tien_tm), float(tien_ck), float(tien_no), item.get('nha_may', '')
        ))
        
    conn.commit()
    conn.close()

def lay_bao_cao_ngay(ngay_str, chi_nhanh="Tất cả", ten_nv="Tất cả"):
    conn = sqlite3.connect(DB_FILE)
    query = "SELECT * FROM don_hang WHERE DATE(ngay_tao) = ?"
    params = [ngay_str]

    if chi_nhanh != "Tất cả":
        query += " AND chi_nhanh = ?"
        params.append(chi_nhanh)

    if ten_nv != "Tất cả":
        query += " AND ten_nv = ?"
        params.append(ten_nv)

    df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    return df

def xoa_don_hang_by_so_don(so_don_hang):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM don_hang WHERE so_don_hang = ?", (so_don_hang,))
    conn.commit()
    conn.close()

def xuat_excel_misa_chuan(ngay_str, chi_nhanh="Tất cả"):
    df = lay_bao_cao_ngay(ngay_str, chi_nhanh, "Tất cả")
    if df.empty:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MISA_Export"

    headers = [
        "STT", "Số đơn hàng", "Ngày", "Chi nhánh", "Nhân viên", "Mã KH", 
        "Tên khách hàng", "Tên người mua", "Mã số thuế", "Địa chỉ", 
        "Mã VT", "Tên VT", "Số lượng", "Đơn giá", "Thành tiền", 
        "Hình thức TT", "Xuất HD", "Tiền TM", "Tiền CK", "Tiền nợ", "Nhà máy"
    ]
    
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    num_cols_indices = [13, 14, 15, 18, 19, 20]

    for idx, row in df.iterrows():
        try:
            ngay_fmt = datetime.strptime(str(row['ngay_tao']), '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y')
        except:
            ngay_fmt = str(row['ngay_tao'])

        sl_val = float(row.get('so_luong', 0))
        dg_val = float(row.get('don_gia', 0))
        tt_val = float(row.get('thanh_tien', 0))
        tm_val = float(row.get('tien_tm', 0))
        ck_val = float(row.get('tien_ck', 0))
        no_val = float(row.get('tien_no', 0))

        raw_ten_kh = str(row.get('ten_kh', '') or '').strip()
        raw_ten_nguoi_mua = str(row.get('ten_nguoi_mua', '') or '').strip()

        clean_ten_kh = "" if raw_ten_kh.lower() == "khách lẻ" else raw_ten_kh
        clean_ten_nguoi_mua = "" if raw_ten_nguoi_mua.lower() == "khách lẻ" else raw_ten_nguoi_mua

        raw_dia_chi = str(row.get('dia_chi', '') or '').strip()
        clean_dia_chi = raw_dia_chi if raw_dia_chi else "Người mua không cung cấp địa chỉ"

        row_data = [
            idx + 1,
            str(row.get('so_don_hang', '')),
            ngay_fmt,
            str(row.get('chi_nhanh', '')),
            str(row.get('ten_nv', '')),
            str(row.get('ma_kh', '')),
            clean_ten_kh,        
            clean_ten_nguoi_mua, 
            str(row.get('ma_so_thue', '')),
            clean_dia_chi,       
            str(row.get('ma_vt', '')),
            str(row.get('ten_vt', '')),
            sl_val,
            dg_val,
            tt_val,
            str(row.get('hinh_thuc_tt', '')),
            str(row.get('xuat_hd', '')),
            tm_val,
            ck_val,
            no_val,
            str(row.get('nha_may', ''))
        ]
        
        ws.append(row_data)
        current_row = ws.max_row

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = Font(name="Arial", size=10)
            
            if col_idx in num_cols_indices:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()